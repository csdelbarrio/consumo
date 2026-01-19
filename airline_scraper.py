#!/usr/bin/env python3
"""
Script de scraping de aerolíneas para detectar personalización de precios.
Simula múltiples usuarios y recopila datos de precios de vuelos y servicios auxiliares.
"""

import time
import random
import json
from datetime import datetime, timedelta
from typing import List, Dict, Optional
import logging
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('airline_scraper.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class UserProfile:
    """Representa un perfil de usuario simulado con diferentes características"""

    USER_AGENTS = [
        # Chrome en Windows
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        # Firefox en Windows
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
        # Safari en Mac
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 14_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15',
        # Chrome en Mac
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        # Chrome en Linux
        'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        # Edge
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36 Edg/120.0.0.0',
    ]

    SCREEN_RESOLUTIONS = [
        (1920, 1080),
        (1366, 768),
        (1440, 900),
        (1536, 864),
        (2560, 1440),
    ]

    def __init__(self, profile_id: int):
        self.profile_id = profile_id
        self.user_agent = random.choice(self.USER_AGENTS)
        self.screen_resolution = random.choice(self.SCREEN_RESOLUTIONS)
        self.cookies_file = Path(f'cookies_profile_{profile_id}.json')

    def get_chrome_options(self) -> Options:
        """Configura las opciones de Chrome para este perfil"""
        options = Options()
        options.add_argument(f'user-agent={self.user_agent}')
        options.add_argument(f'--window-size={self.screen_resolution[0]},{self.screen_resolution[1]}')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)

        # Headless opcional (descomenta si quieres que no se vea la ventana)
        # options.add_argument('--headless')

        # Opciones para evitar detección
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--no-sandbox')

        return options

    def save_cookies(self, driver: webdriver.Chrome):
        """Guarda las cookies del navegador"""
        cookies = driver.get_cookies()
        with open(self.cookies_file, 'w') as f:
            json.dump(cookies, f)
        logger.info(f"Cookies guardadas para perfil {self.profile_id}")

    def load_cookies(self, driver: webdriver.Chrome):
        """Carga las cookies guardadas"""
        if self.cookies_file.exists():
            with open(self.cookies_file, 'r') as f:
                cookies = json.load(f)
            for cookie in cookies:
                try:
                    driver.add_cookie(cookie)
                except Exception as e:
                    logger.warning(f"No se pudo cargar cookie: {e}")
            logger.info(f"Cookies cargadas para perfil {self.profile_id}")


class AirlineScraper:
    """Scraper principal para aerolíneas"""

    def __init__(self, profile: UserProfile):
        self.profile = profile
        self.driver = None
        self.results = []

    def __enter__(self):
        self.driver = webdriver.Chrome(options=self.profile.get_chrome_options())
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.driver:
            self.driver.quit()

    def random_delay(self, min_seconds: float = 2, max_seconds: float = 5):
        """Espera aleatoria para simular comportamiento humano"""
        time.sleep(random.uniform(min_seconds, max_seconds))

    def scrape_iberia(self, origin: str, destination: str, date: str) -> Dict:
        """
        Scrape de Iberia.com

        Args:
            origin: Código aeropuerto origen (ej: 'MAD')
            destination: Código aeropuerto destino (ej: 'BCN')
            date: Fecha en formato 'YYYY-MM-DD'
        """
        logger.info(f"Scraping Iberia: {origin} -> {destination} el {date}")

        result = {
            'timestamp': datetime.now().isoformat(),
            'profile_id': self.profile.profile_id,
            'airline': 'Iberia',
            'origin': origin,
            'destination': destination,
            'date': date,
            'user_agent': self.profile.user_agent,
            'screen_resolution': f"{self.profile.screen_resolution[0]}x{self.profile.screen_resolution[1]}",
        }

        try:
            # Construir URL de Iberia
            url = f"https://www.iberia.com/es/"
            self.driver.get(url)
            self.random_delay(3, 5)

            # Aquí iría la lógica específica de scraping para Iberia
            # NOTA: Esto es un ejemplo básico, necesitas ajustarlo a la estructura actual de la web

            # Ejemplo de búsqueda de precios (adaptar a la estructura real)
            try:
                # Esperar a que cargue la página
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )

                # Aquí buscarías los elementos de precio
                # Por ejemplo: price_elements = self.driver.find_elements(By.CLASS_NAME, "price")

                result['basic_price'] = None  # Extraer precio básico
                result['checked_bag'] = None  # Precio maleta facturada
                result['seat_selection'] = None  # Precio selección asiento
                result['priority_boarding'] = None  # Precio embarque prioritario
                result['status'] = 'success'
                result['error'] = None

            except TimeoutException:
                result['status'] = 'timeout'
                result['error'] = 'Timeout esperando elementos de la página'
                logger.warning(f"Timeout en Iberia para {origin}->{destination}")

        except Exception as e:
            result['status'] = 'error'
            result['error'] = str(e)
            logger.error(f"Error scraping Iberia: {e}")

        return result

    def scrape_ryanair(self, origin: str, destination: str, date: str) -> Dict:
        """
        Scrape de Ryanair.com

        Args:
            origin: Código aeropuerto origen
            destination: Código aeropuerto destino
            date: Fecha en formato 'YYYY-MM-DD'
        """
        logger.info(f"Scraping Ryanair: {origin} -> {destination} el {date}")

        result = {
            'timestamp': datetime.now().isoformat(),
            'profile_id': self.profile.profile_id,
            'airline': 'Ryanair',
            'origin': origin,
            'destination': destination,
            'date': date,
            'user_agent': self.profile.user_agent,
            'screen_resolution': f"{self.profile.screen_resolution[0]}x{self.profile.screen_resolution[1]}",
        }

        try:
            # URL de búsqueda de Ryanair
            url = "https://www.ryanair.com/es/es"
            self.driver.get(url)
            self.random_delay(3, 5)

            # Lógica específica para Ryanair
            # NOTA: Adaptar a la estructura actual de la web

            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )

                result['basic_price'] = None
                result['checked_bag'] = None
                result['seat_selection'] = None
                result['priority_boarding'] = None
                result['status'] = 'success'
                result['error'] = None

            except TimeoutException:
                result['status'] = 'timeout'
                result['error'] = 'Timeout esperando elementos de la página'
                logger.warning(f"Timeout en Ryanair para {origin}->{destination}")

        except Exception as e:
            result['status'] = 'error'
            result['error'] = str(e)
            logger.error(f"Error scraping Ryanair: {e}")

        return result

    def scrape_vueling(self, origin: str, destination: str, date: str) -> Dict:
        """
        Scrape de Vueling.com

        Args:
            origin: Código aeropuerto origen
            destination: Código aeropuerto destino
            date: Fecha en formato 'YYYY-MM-DD'
        """
        logger.info(f"Scraping Vueling: {origin} -> {destination} el {date}")

        result = {
            'timestamp': datetime.now().isoformat(),
            'profile_id': self.profile.profile_id,
            'airline': 'Vueling',
            'origin': origin,
            'destination': destination,
            'date': date,
            'user_agent': self.profile.user_agent,
            'screen_resolution': f"{self.profile.screen_resolution[0]}x{self.profile.screen_resolution[1]}",
        }

        try:
            url = "https://www.vueling.com/es"
            self.driver.get(url)
            self.random_delay(3, 5)

            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.TAG_NAME, "body"))
                )

                result['basic_price'] = None
                result['checked_bag'] = None
                result['seat_selection'] = None
                result['priority_boarding'] = None
                result['status'] = 'success'
                result['error'] = None

            except TimeoutException:
                result['status'] = 'timeout'
                result['error'] = 'Timeout esperando elementos de la página'
                logger.warning(f"Timeout en Vueling para {origin}->{destination}")

        except Exception as e:
            result['status'] = 'error'
            result['error'] = str(e)
            logger.error(f"Error scraping Vueling: {e}")

        return result


class PriceAnalyzer:
    """Analiza los resultados y detecta personalización de precios"""

    def __init__(self, results: List[Dict]):
        self.results = results
        self.df = pd.DataFrame(results)

    def save_to_excel(self, filename: str = None):
        """Guarda los resultados en un archivo Excel"""
        if filename is None:
            filename = f"airline_prices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Crear archivo Excel con formato
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Hoja de datos crudos
            self.df.to_excel(writer, sheet_name='Datos', index=False)

            # Hoja de análisis comparativo
            if len(self.df) > 0:
                analysis_df = self._create_analysis_dataframe()
                analysis_df.to_excel(writer, sheet_name='Análisis', index=False)

            # Aplicar formato
            workbook = writer.book

            # Formato para la hoja de datos
            worksheet_data = writer.sheets['Datos']
            self._format_worksheet(worksheet_data)

            # Formato para la hoja de análisis
            if 'Análisis' in writer.sheets:
                worksheet_analysis = writer.sheets['Análisis']
                self._format_worksheet(worksheet_analysis)

        logger.info(f"Resultados guardados en {filename}")
        return filename

    def _create_analysis_dataframe(self) -> pd.DataFrame:
        """Crea un DataFrame con análisis comparativo de precios"""
        analysis = []

        # Agrupar por ruta, fecha y aerolínea
        for (airline, origin, destination, date), group in self.df.groupby(
            ['airline', 'origin', 'destination', 'date']
        ):
            if len(group) > 1:  # Solo si hay múltiples perfiles
                prices = group['basic_price'].dropna()
                if len(prices) > 1:
                    analysis.append({
                        'airline': airline,
                        'route': f"{origin} -> {destination}",
                        'date': date,
                        'profiles_count': len(group),
                        'min_price': prices.min(),
                        'max_price': prices.max(),
                        'avg_price': prices.mean(),
                        'price_difference': prices.max() - prices.min(),
                        'price_variation_%': ((prices.max() - prices.min()) / prices.min() * 100) if prices.min() > 0 else 0,
                        'possible_personalization': 'SÍ' if (prices.max() - prices.min()) > 0 else 'NO'
                    })

        return pd.DataFrame(analysis)

    def _format_worksheet(self, worksheet):
        """Aplica formato a una hoja de Excel"""
        # Formato de encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def print_summary(self):
        """Imprime un resumen de los resultados"""
        print("\n" + "="*80)
        print("RESUMEN DE SCRAPING DE AEROLÍNEAS")
        print("="*80)

        if len(self.df) == 0:
            print("No hay datos para mostrar")
            return

        print(f"\nTotal de búsquedas: {len(self.df)}")
        print(f"Búsquedas exitosas: {len(self.df[self.df['status'] == 'success'])}")
        print(f"Búsquedas con error: {len(self.df[self.df['status'] != 'success'])}")

        print("\nAerolíneas analizadas:")
        for airline, count in self.df['airline'].value_counts().items():
            print(f"  - {airline}: {count} búsquedas")

        print("\nPerfiles de usuario utilizados:")
        for profile_id, count in self.df['profile_id'].value_counts().items():
            print(f"  - Perfil {profile_id}: {count} búsquedas")

        # Análisis de personalización
        analysis_df = self._create_analysis_dataframe()
        if len(analysis_df) > 0:
            personalization_detected = analysis_df[analysis_df['possible_personalization'] == 'SÍ']
            if len(personalization_detected) > 0:
                print("\n⚠️  POSIBLE PERSONALIZACIÓN DE PRECIOS DETECTADA:")
                for _, row in personalization_detected.iterrows():
                    print(f"  - {row['airline']} {row['route']} ({row['date']}): "
                          f"diferencia de {row['price_difference']:.2f}€ "
                          f"({row['price_variation_%']:.1f}%)")
            else:
                print("\n✓ No se detectó personalización de precios significativa")

        print("\n" + "="*80 + "\n")


def run_scraping_session(routes: List[Dict], num_profiles: int = 3) -> List[Dict]:
    """
    Ejecuta una sesión completa de scraping

    Args:
        routes: Lista de diccionarios con 'airline', 'origin', 'destination', 'date'
        num_profiles: Número de perfiles de usuario a simular

    Returns:
        Lista de resultados
    """
    all_results = []

    for profile_id in range(1, num_profiles + 1):
        logger.info(f"\n{'='*60}")
        logger.info(f"Iniciando scraping con Perfil {profile_id}")
        logger.info(f"{'='*60}")

        profile = UserProfile(profile_id)

        with AirlineScraper(profile) as scraper:
            for route in routes:
                try:
                    airline = route['airline'].lower()

                    if airline == 'iberia':
                        result = scraper.scrape_iberia(
                            route['origin'],
                            route['destination'],
                            route['date']
                        )
                    elif airline == 'ryanair':
                        result = scraper.scrape_ryanair(
                            route['origin'],
                            route['destination'],
                            route['date']
                        )
                    elif airline == 'vueling':
                        result = scraper.scrape_vueling(
                            route['origin'],
                            route['destination'],
                            route['date']
                        )
                    else:
                        logger.warning(f"Aerolínea no soportada: {airline}")
                        continue

                    all_results.append(result)

                    # Delay entre búsquedas
                    time.sleep(random.uniform(10, 20))

                except Exception as e:
                    logger.error(f"Error en ruta {route}: {e}")
                    continue

        # Delay entre perfiles
        if profile_id < num_profiles:
            logger.info(f"Esperando antes del siguiente perfil...")
            time.sleep(random.uniform(30, 60))

    return all_results


def main():
    """Función principal"""
    # Configuración de rutas a scrapear
    # Fecha de ejemplo: 30 días desde hoy
    search_date = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')

    routes = [
        # Puedes agregar más rutas aquí
        {'airline': 'Iberia', 'origin': 'MAD', 'destination': 'BCN', 'date': search_date},
        {'airline': 'Ryanair', 'origin': 'MAD', 'destination': 'PMI', 'date': search_date},
        {'airline': 'Vueling', 'origin': 'BCN', 'destination': 'MAD', 'date': search_date},
    ]

    logger.info("="*80)
    logger.info("INICIANDO BOT DE SCRAPING DE AEROLÍNEAS")
    logger.info("="*80)
    logger.info(f"Fecha de búsqueda: {search_date}")
    logger.info(f"Rutas a analizar: {len(routes)}")
    logger.info(f"Perfiles a utilizar: 3")

    # Ejecutar scraping
    results = run_scraping_session(routes, num_profiles=3)

    # Analizar resultados
    analyzer = PriceAnalyzer(results)
    analyzer.print_summary()

    # Guardar en Excel
    excel_file = analyzer.save_to_excel()
    logger.info(f"\n✓ Proceso completado. Resultados guardados en: {excel_file}")


if __name__ == "__main__":
    main()
